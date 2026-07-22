"""Export formatters — XLSX, CSV, PDF.

Each formatter takes a list of Dataset objects and returns bytes.
"""

from __future__ import annotations

import csv
import io
from datetime import datetime

from .data import Dataset

# ---------------------------------------------------------------------------
# XLSX (openpyxl)
# ---------------------------------------------------------------------------


def to_xlsx(datasets: list[Dataset], surface_label: str = "Export") -> bytes:
    """Generate an .xlsx workbook with one sheet per dataset."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl is required for XLSX export: pip install openpyxl")

    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3C4A", end_color="1A3C4A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )
    num_fmt_money = '#,##0'
    num_fmt_pct = '0.0%'

    for ds in datasets:
        title = ds.title[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=title)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(ds.columns), 1))
        title_cell = ws.cell(row=1, column=1, value=f"{surface_label} — {ds.title}")
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="1A3C4A")
        ws.cell(row=2, column=1, value=f"Exported {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color="888888")

        # Header row (row 4)
        header_row = 4
        for col_idx, col_name in enumerate(ds.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Data rows
        for row_idx, row in enumerate(ds.rows, header_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                # Right-align numbers
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = num_fmt_money

        # Auto-width columns (approximate)
        from openpyxl.utils import get_column_letter

        for col_idx, col_name in enumerate(ds.columns, 1):
            max_len = len(str(col_name))
            for row in ds.rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # If no datasets produced sheets, add an empty one
    if not wb.sheetnames:
        wb.create_sheet("No Data")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def to_csv(datasets: list[Dataset], surface_label: str = "Export") -> bytes:
    """Generate CSV bytes. Multiple datasets are separated by blank lines."""
    buf = io.StringIO()
    writer = csv.writer(buf)

    for i, ds in enumerate(datasets):
        if i > 0:
            writer.writerow([])  # blank separator

        writer.writerow([f"# {surface_label} — {ds.title}"])
        writer.writerow([f"# Exported {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        writer.writerow(ds.columns)
        for row in ds.rows:
            writer.writerow(row)

    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# PDF (reportlab)
# ---------------------------------------------------------------------------


def to_pdf(datasets: list[Dataset], surface_label: str = "Export") -> bytes:
    """Generate a styled PDF report with tables for each dataset."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
    except ImportError:
        raise ImportError("reportlab is required for PDF export: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(letter),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.6 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ApexTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1A3C4A"),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "ApexSubtitle",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#888888"),
        spaceAfter=14,
    )
    section_style = ParagraphStyle(
        "ApexSection",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=colors.HexColor("#0C998D"),
        spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        "ApexCell",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
    )
    header_cell_style = ParagraphStyle(
        "ApexHeaderCell",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        textColor=colors.white,
    )

    elements: list = []

    # Report title
    elements.append(Paragraph(f"APEX — {surface_label}", title_style))
    elements.append(
        Paragraph(
            f"Exported {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            subtitle_style,
        )
    )

    for ds in datasets:
        elements.append(Paragraph(ds.title, section_style))

        if not ds.columns and not ds.rows:
            elements.append(Paragraph("No data available.", styles["Normal"]))
            elements.append(Spacer(1, 12))
            continue

        # Build table data
        table_data = []
        # Header
        header = [Paragraph(f"<b>{c}</b>", header_cell_style) for c in ds.columns]
        table_data.append(header)
        # Data rows
        for row in ds.rows:
            table_data.append(
                [Paragraph(str(v) if v is not None else "", cell_style) for v in row]
            )

        # Calculate column widths
        available_width = landscape(letter)[0] - 1.0 * inch
        n_cols = len(ds.columns) or 1
        col_width = available_width / n_cols

        table = Table(table_data, colWidths=[col_width] * n_cols, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3C4A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FA")]),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        elements.append(table)
        elements.append(Spacer(1, 16))

    doc.build(elements)
    return buf.getvalue()
